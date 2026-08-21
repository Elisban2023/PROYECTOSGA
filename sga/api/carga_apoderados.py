from io import BytesIO

from django.db import transaction
from django.http import FileResponse
from drf_spectacular.utils import OpenApiTypes, extend_schema
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from sga.models import RegistroAuditoria
from sga.permissions import IsAdminOrDirectivo
from sga.serializers.usuarios import ApoderadoSerializer

ENCABEZADOS = (
    "username", "password", "first_name", "last_name", "email", "dni",
    "telefono", "is_active",
)


def _cell_value(value):
    if value is None:
        return ""
    return value.isoformat() if hasattr(value, "isoformat") else str(value).strip()


@extend_schema(responses=OpenApiTypes.BINARY)
@api_view(["GET"])
@permission_classes([IsAdminOrDirectivo])
def plantilla_carga_apoderados(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Apoderados"
    sheet.append(ENCABEZADOS)
    sheet.append((
        "apoderado.ejemplo", "Temporal2026!SGA", "Nombres", "Apellidos",
        "apoderado@example.com", "12345678", "999999999", "true",
    ))
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 22
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return FileResponse(
        output,
        as_attachment=True,
        filename="plantilla_carga_apoderados.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@extend_schema(request=OpenApiTypes.OBJECT, responses=OpenApiTypes.OBJECT)
@api_view(["POST"])
@permission_classes([IsAdminOrDirectivo])
def carga_masiva_apoderados(request):
    archivo = request.FILES.get("archivo")
    if archivo is None:
        raise ValidationError({"archivo": "Debe adjuntar un archivo Excel .xlsx."})
    if not archivo.name.lower().endswith(".xlsx"):
        raise ValidationError({"archivo": "El archivo debe tener formato .xlsx."})
    if archivo.size > 5 * 1024 * 1024:
        raise ValidationError({"archivo": "El archivo no puede superar 5 MB."})
    try:
        rows = list(load_workbook(archivo, read_only=True, data_only=True).active.iter_rows(values_only=True))
    except Exception as exc:
        raise ValidationError({"archivo": "No se pudo leer el archivo Excel."}) from exc
    if not rows:
        raise ValidationError({"archivo": "El archivo no contiene filas."})
    headers = tuple(_cell_value(value) for value in rows[0])
    if headers != ENCABEZADOS:
        raise ValidationError({"archivo": {"encabezados_esperados": ENCABEZADOS}})
    if len(rows) == 1:
        raise ValidationError({"archivo": "Agregue al menos un apoderado."})
    datos, errores, usernames, dnis = [], [], set(), set()
    for number, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        payload = dict(zip(ENCABEZADOS, (_cell_value(value) for value in row)))
        if payload["username"].lower() in usernames:
            errores.append({"fila": number, "errores": {"username": ["Duplicado dentro del archivo."]}})
            continue
        if payload["dni"] in dnis:
            errores.append({"fila": number, "errores": {"dni": ["Duplicado dentro del archivo."]}})
            continue
        usernames.add(payload["username"].lower())
        dnis.add(payload["dni"])
        serializer = ApoderadoSerializer(data=payload)
        if not serializer.is_valid():
            errores.append({"fila": number, "errores": serializer.errors})
        else:
            datos.append(serializer)
    if errores:
        return Response({"creados": 0, "errores": errores}, status=status.HTTP_400_BAD_REQUEST)
    with transaction.atomic():
        apoderados = [serializer.save() for serializer in datos]
        RegistroAuditoria.registrar_evento(
            user=request.user,
            accion="CARGA_MASIVA_APODERADOS",
            modulo="usuarios",
            entidad="Apoderado",
            entidad_id=str(len(apoderados)),
        )
    return Response({"creados": len(apoderados), "apoderados": ApoderadoSerializer(apoderados, many=True).data}, status=status.HTTP_201_CREATED)
